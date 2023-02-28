import os
import openpyxl

# Get the path to the folder
path = r"D:\Academic\MSc\5th Semester\Project\pythonProject\Results"

# Create a new workbook
wb = openpyxl.Workbook()

# Select the active worksheet
ws = wb.active

# Write the header row
ws['A1'] = "Name"
ws['B1'] = "Image Name"
ws['C1'] = "final_images_ok"
ws['D1'] = "final_images_not"

# Loop through the folders
row = 2
for name in os.listdir(path):
    if not os.path.isdir(os.path.join(path, name)):
        continue

    # Get the path to the "final_images_ok" and "final_images_not" folders
    ok_path = os.path.join(path, name, "final_images_ok")
    not_path = os.path.join(path, name, "final_images_not")

    # Count the number of files in each folder
    ok_files = os.listdir(ok_path)
    not_files = os.listdir(not_path)

    # Write the data to the worksheet
    ws.cell(row=row, column=1, value=name)
    if len(ok_files) > 0:
        image_name = os.path.splitext(ok_files[0])[0]
        image_name = "_".join(image_name.split("_")[:-1])
        ws.cell(row=row, column=2, value=image_name)
    elif len(not_files) > 0:
        image_name = os.path.splitext(not_files[0])[0]
        image_name = "_".join(image_name.split("_")[:-1])
        ws.cell(row=row, column=2, value=image_name)
    ws.cell(row=row, column=3, value=len(ok_files))
    ws.cell(row=row, column=4, value=len(not_files))
    row += 1

# Save the workbook
wb.save("result.xlsx")
