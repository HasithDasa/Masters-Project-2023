from json import load
import numpy as np
import cv2
# import os
from timeit import default_timer as timer
import concurrent.futures



start_time_prep = timer()

def rle_decode(mask_rle, shape):
    s = mask_rle
    starts, lengths = [np.asarray(x, dtype=int) for x in (s[0:][::2], s[1:][::2])]
    starts -= 1
    ends = starts + lengths
    shape = shape[1], shape[0]
    img = np.zeros(shape[0] * shape[1], dtype=np.uint8)
    # print(img)
    for lo, hi in zip(starts, ends):
        img[lo:hi] = 1
    return img.reshape(shape)


with open('training_done.json') as f:
    data = load(f)
    img_no = 8
    data1 = data['images'][img_no]  # selecting the image
    img_name = data1['image_name']
    img_name = img_name.replace('.png', '')
    img_width = data1['width']
    img_height = data1['height']
    img_temp_shape = (img_height, img_width)
    img_temp = np.zeros(img_temp_shape, dtype=np.uint8)
    img_temp_3 = np.zeros(img_temp_shape, dtype=np.uint8)
    img_temp_5 = np.zeros(img_temp_shape, dtype=np.uint8)
    img_temp_4 = np.zeros(img_temp_shape, dtype=np.uint8)
    img_temp_8 = np.zeros(img_temp_shape, dtype=np.uint8)
    data2 = data1['labels']  # selecting the label key


    list_of_keys = []
    list_of_keys_t5 = []
    dic_for_selection = {}
    dic_for_selection_t4 = {}
    dic_for_selection_t5 = {}


    bbox_list = []
    lg_bbox_list =[]
    # co_with_ones = []
    bbox_ones_mask_list = []
    mask_list = []
    intersected_comm_ones_list = []
    save_comm_list_1 = []
    save_comm_list_2 = []

    # slope_angles_for_classification= []

# function => converting to opencv format (def opencv_format)

    for i in range(len(data2)):
        data3 = data2[i]  # selecting the class label
        mask_rle = data3['mask']
        bbox = data3['bbox']
        class_name = data3['class_name']


        if class_name == 'First Section Cutting' or class_name == 'Redundant Top End' or class_name == 'Redundant Bottom End' or class_name == 'Tip Cutting' or class_name == 'Non-Viable Part' or class_name == 'Second Section Cutting' or class_name == 'Third Section Cutting' or class_name == 'Fourth Section Cutting':
            # making mask
            mask = rle_decode(mask_rle, (bbox[2] - bbox[0], bbox[3] - bbox[1]))

            # where exactly the mask situated in main image
            row_start = bbox[1]
            row_end = bbox[3]

            col_start = bbox[0]
            col_end = bbox[2]

            #collecting bbox information
            bbox_list.append([col_start, col_end, row_start, row_end])

            img_temp[row_start:row_end, col_start:col_end] = mask

            variable = str(i)
            key = "lable" + variable
            list_of_keys.append(key)
            dic_for_selection[key] = img_temp

            # including mask in the main image, need to make it zero for everytime because of mask coincidence problem
            img_temp = np.zeros(img_temp_shape, dtype=np.uint8)

        # considering large bounding boxes to eliminate unwanted intersections of small bounding boxes
        if class_name == "Raw Cutting":
            # where exactly the mask situated in main image
            lg_row_start = bbox[1]
            lg_row_end = bbox[3]

            lg_col_start = bbox[0]
            lg_col_end = bbox[2]

            # collecting large bbox information
            lg_bbox_list.append([lg_col_start, lg_col_end, lg_row_start, lg_row_end])

    # print("lg_bbox_list", lg_bbox_list)

    edge_only_list =[]

    for j in range(len(list_of_keys)):
        # print(list_of_keys)

        # img_temp_3 = np.zeros(img_temp_shape, dtype=np.uint8)

        class_name_from_list = list_of_keys[j]
        selected_list = dic_for_selection[class_name_from_list]
        mask_list.append(selected_list)
        # co_with_ones.append(selected_list)

        result_with_one = np.where(selected_list == 1)
        listOfCoordinates = list(zip(result_with_one[0], result_with_one[1]))

        # co_with_ones.append(listOfCoordinates)

        for cord in listOfCoordinates:
            img_temp_3[cord] = 1

        #img_contour_only = img_temp_3.astype(np.uint8) * 255


        contours_bb, hierarchy_bb = cv2.findContours(img_temp_3, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        co_with_ones_only_edge_elem_coord = []


        for co_with_ones_only_edge_elem in contours_bb:

            for point in co_with_ones_only_edge_elem:
                y, x = point[0]
                co_with_ones_only_edge_elem_coord.append((x, y))
            # co_with_ones_only_edge.append(co_with_ones_only_edge_elem_coord)



        edge_only_list.append(co_with_ones_only_edge_elem_coord)

        img_temp_3 = np.zeros(img_temp_shape, dtype=np.uint8)


    def my_function(bbox_list, lg_bbox_list, edge_only_list):
        with concurrent.futures.ThreadPoolExecutor() as executor:
            # Submit the first for loop as a task
            task1 = executor.submit(process_first_for_loop, bbox_list, edge_only_list)

            # Submit the second for loop as a task
            task2 = executor.submit(process_second_for_loop, lg_bbox_list, edge_only_list)

            # Wait for both tasks to complete
            results = [task1.result(), task2.result()]

        # Combine the results from both tasks and return them
        return results


    def process_first_for_loop(bbox_list, edge_only_list):
        bbox_ones_mask_list = []
        intersected_comm_ones_list = []


        for mask_index in range(len(bbox_list)):
            for mask_index_checked in range(len(edge_only_list)):
                for ones_index in range(len(edge_only_list[mask_index_checked])):
                    if (bbox_list[mask_index][0] <= edge_only_list[mask_index_checked][ones_index][1] < bbox_list[mask_index][1]) and (bbox_list[mask_index][2] <= edge_only_list[mask_index_checked][ones_index][0] < bbox_list[mask_index][3]):
                        if mask_index != mask_index_checked:
                            bbox_ones_mask_list.append([mask_index, mask_index_checked])
                            intersected_comm_ones_list.append([(edge_only_list[mask_index_checked][ones_index][0],
                                                                edge_only_list[mask_index_checked][ones_index][1]),
                                                               mask_index, mask_index_checked])
        return (bbox_ones_mask_list, intersected_comm_ones_list) # remove brackets for non parallel


    def process_second_for_loop(lg_bbox_list, edge_only_list):

        mask_inside_lg_bbox_list = []
        for lg_bb_ind, lg_bb_ele in enumerate(lg_bbox_list):
            for mask_index_2, mask_ele_2 in enumerate(edge_only_list):
                ones_count = 0
                for co_with_ones_ele2 in mask_ele_2:
                    if (lg_bb_ele[0] <= co_with_ones_ele2[1] < lg_bb_ele[1]) and (
                            lg_bb_ele[2] <= co_with_ones_ele2[0] < lg_bb_ele[3]):
                        ones_count += 1
                if ones_count == len(mask_ele_2):
                    mask_inside_lg_bbox_list.append([lg_bb_ind, mask_index_2])
        return mask_inside_lg_bbox_list


    results = my_function(bbox_list, lg_bbox_list, edge_only_list)

    bbox_ones_mask_list = results[0][0]
    intersected_comm_ones_list = results[0][1]
    mask_inside_lg_bbox_list = results[1]

    # mask_inside_lg_bbox_list = process_second_for_loop(lg_bbox_list, co_with_ones)
    # bbox_ones_mask_list, intersected_comm_ones_list = process_first_for_loop(bbox_list, co_with_ones)



    unique_sublists = set()

    for sublist in bbox_ones_mask_list:
        if sublist[0] < sublist[1]:
            unique_sublists.add(tuple(sublist))
        else:
            unique_sublists.add(tuple(reversed(sublist)))

    # Convert the set back to a list of lists
    intersected_masks_list = [list(sublist) for sublist in unique_sublists]

    # Group sublists by their first element
    grouped_list = {}


    for sublist in mask_inside_lg_bbox_list:
        key = sublist[0]
        if key not in grouped_list:
            grouped_list[key] = []
        grouped_list[key].append(sublist)

    # Create new list with second elements of sublists in each group
    intersected_bbox_in_bigbb_list = []
    for group in grouped_list.values():
        new_sublist = [sublist[1] for sublist in group]
        intersected_bbox_in_bigbb_list.append(new_sublist)

    # Create a set of all unique elements in intersected_bbox_in_bigbb_list
    all_elements = set()
    for sublist in intersected_bbox_in_bigbb_list:
        all_elements.update(sublist)

    intersected_masks_list_final = []

    for mask in intersected_masks_list:
        found = False
        for bbox in intersected_bbox_in_bigbb_list:
            if all(elem in bbox for elem in mask):
                intersected_masks_list_final.append(mask)
                found = True
                break
        if not found:
            continue


    # print('intersected_bbox_in_bigbb_list', intersected_bbox_in_bigbb_list)
    # print("intersected_masks_list", intersected_masks_list)
    # print("intersected_masks_list_final", intersected_masks_list_final)

# function => releasing image unit (def rel_img_unit)

    #going on one by one on intersected masks

    for ele_ind, ele_int_list in enumerate(intersected_masks_list_final):

        # print("ele_int_list", ele_int_list)

        for save_comm in intersected_comm_ones_list:
            # print("save_comm", save_comm[1], save_comm[2])
            if save_comm[1] == ele_int_list[0] and save_comm[2] == ele_int_list[1]: # common parts in 2nd image (in 1st image)
                save_comm_list_1.append(save_comm[0])

            if save_comm[1] == ele_int_list[1] and save_comm[2] == ele_int_list[0]: # common parts in 2nd image (in 1st image)
                save_comm_list_2.append(save_comm[0])

        for coord_comm_1 in save_comm_list_1:
            img_temp_5[coord_comm_1] = 1

        for coord_comm_2 in save_comm_list_2:
            img_temp_4[coord_comm_2] = 1

        variable_t5 = str(ele_ind)
        key_t5 = "t5" + variable_t5
        list_of_keys_t5.append(key_t5)

        dic_for_selection_t5[key_t5] = img_temp_5
        dic_for_selection_t4[key_t5] = img_temp_4

        # making image masks zero after iteration
        img_temp_5 = np.zeros(img_temp_shape, dtype=np.uint8)
        img_temp_4 = np.zeros(img_temp_shape, dtype=np.uint8)

    end_time_prep = timer()
    elapsed_time_prep = (end_time_prep - start_time_prep) * 1000
    print("elapsed_time_prep", elapsed_time_prep)

# function => seperation of line segments from the image(def acq_line_seg)
    start_time_sing_line = timer()
    to_time_excel = []

    for list_of_keys_t5_ind in range(len(list_of_keys_t5)):

        class_name_from_list_of_keys_t5 = list_of_keys_t5[list_of_keys_t5_ind]
        selected_list_t5 = dic_for_selection_t5[class_name_from_list_of_keys_t5]
        selected_list_t4 = dic_for_selection_t4[class_name_from_list_of_keys_t5]
        print("image sequence ID:", list_of_keys_t5_ind)

#  Canny Try
        if list_of_keys_t5_ind > 0:
            inverse_selected_list_t5 = cv2.bitwise_not(dic_for_selection_t5[list_of_keys_t5[list_of_keys_t5_ind - 1]])
            selected_list_t5 = (cv2.bitwise_and(selected_list_t5, inverse_selected_list_t5)).astype(np.uint8) * 255

            inverse_selected_list_t4 = cv2.bitwise_not(dic_for_selection_t4[list_of_keys_t5[list_of_keys_t5_ind - 1]])
            selected_list_t4 = (cv2.bitwise_and(selected_list_t4, inverse_selected_list_t4)).astype(np.uint8) * 255

        else:
            selected_list_t5 = selected_list_t5.astype(np.uint8) * 255
            selected_list_t4 = selected_list_t4.astype(np.uint8) * 255


        ## canny try
        edges_canny_temp_5 = cv2.Canny(selected_list_t5, 1, 250)
        edges_canny_temp_4 = cv2.Canny(selected_list_t4, 1, 250)

        and_edges_t4_t5 = cv2.bitwise_and(edges_canny_temp_5, edges_canny_temp_4)


        row_indexes, col_indexes = np.nonzero(and_edges_t4_t5)

        # Sobel Try

        if len(row_indexes) == 0:

            # print("inside here sobel")

            if list_of_keys_t5_ind > 0:
                inverse_selected_list_t5 = cv2.bitwise_not(dic_for_selection_t5[list_of_keys_t5[list_of_keys_t5_ind - 1]])
                selected_list_t5 = (cv2.bitwise_and(selected_list_t5, inverse_selected_list_t5))

                inverse_selected_list_t4 = cv2.bitwise_not(dic_for_selection_t4[list_of_keys_t5[list_of_keys_t5_ind - 1]])
                selected_list_t4 = (cv2.bitwise_and(selected_list_t4, inverse_selected_list_t4))

            else:
                selected_list_t5 = selected_list_t5
                selected_list_t4 = selected_list_t4

            # Apply Sobel operator t5
            sobelx_t5 = cv2.Sobel(selected_list_t5, cv2.CV_8U, 1, 0, ksize=5)
            sobely_t5 = cv2.Sobel(selected_list_t5, cv2.CV_8U, 0, 1, ksize=5)
            sobel_t5 = sobelx_t5 + sobely_t5

            # Apply Sobel operator t4
            sobelx_t4 = cv2.Sobel(selected_list_t4, cv2.CV_8U, 1, 0, ksize=5)
            sobely_t4 = cv2.Sobel(selected_list_t4, cv2.CV_8U, 0, 1, ksize=5)
            sobel_t4 = sobelx_t4 + sobely_t4

            and_edges_t4_t5 = cv2.bitwise_and(sobel_t5, sobel_t4)

        row_indexes, col_indexes = np.nonzero(and_edges_t4_t5)

        # LoG

        if len(row_indexes) == 0:

            # print("inside here LoG")

            if list_of_keys_t5_ind > 0:
                inverse_selected_list_t5 = cv2.bitwise_not(dic_for_selection_t5[list_of_keys_t5[list_of_keys_t5_ind - 1]])
                selected_list_t5 = (cv2.bitwise_and(selected_list_t5, inverse_selected_list_t5))

                inverse_selected_list_t4 = cv2.bitwise_not(dic_for_selection_t4[list_of_keys_t5[list_of_keys_t5_ind - 1]])
                selected_list_t4 = (cv2.bitwise_and(selected_list_t4, inverse_selected_list_t4))

            else:
                selected_list_t5 = selected_list_t5
                selected_list_t4 = selected_list_t4

            gaussian_image_t5 = cv2.GaussianBlur(selected_list_t5, (3, 3), 0)
            log_t5 = cv2.Laplacian(gaussian_image_t5, cv2.CV_8U)

            gaussian_image_t4 = cv2.GaussianBlur(selected_list_t4, (3, 3), 0)
            log_t4 = cv2.Laplacian(gaussian_image_t4, cv2.CV_8U)

            and_edges_t4_t5 = cv2.bitwise_and(log_t5, log_t4)

            row_indexes, col_indexes = np.nonzero(and_edges_t4_t5)


        # print("row_indexes", row_indexes, "col_indexes", col_indexes)



# function => checking for clustering (def check_clust)


        clustering_list = []
        clustering_len_list = []
        clustering_list_col = []
        clustering_list_row = []

        clustering_needed = False
        check_difference_in_btwn_clus_lst = False
        clustering_type = "row"
        has_duplicates = False
        only_ones_zeroes = False

        for row_ind, row_val in enumerate(row_indexes):
            if row_ind < (len(row_indexes)-1):
                if abs(row_val-row_indexes[row_ind+1]) > 3: # checking the distance between two neighbouring pixels are higher than 3
                    clustering_list.append(row_ind)
                    clustering_type = "row"
                    # print("clustering_type", clustering_type)

        clustering_list_row = clustering_list

        if len(clustering_list) > 0:

            for clustering_list_elem in range(1, len(clustering_list)): # starting from one
                distance_btwn_clus_lst = clustering_list[clustering_list_elem] - clustering_list[clustering_list_elem - 1] if clustering_list_elem < len(clustering_list) - 1 else len(row_indexes) - clustering_list[clustering_list_elem]
                if distance_btwn_clus_lst > 50 :
                    check_difference_in_btwn_clus_lst = True

            if clustering_list[0] > 50: # include 0th element  > 10
                check_difference_in_btwn_clus_lst = True


        if len(clustering_list) == 0 or check_difference_in_btwn_clus_lst == True:

            clustering_list = []

            for col_ind, col_val in enumerate(col_indexes):
                if col_ind < (len(col_indexes) - 1):
                    if abs(col_val-col_indexes[col_ind+1]) > 3:
                        clustering_list.append(col_ind)
                        clustering_type = "col"
                        # print("clustering_type", clustering_type)

            clustering_list_col = clustering_list

        # print("clustering_list", clustering_list)

# function => checking the which clustering it belongs to (def wh_clust)

        if len(clustering_list_col) > 0 or len(clustering_list_row) > 0:
            clustering_needed = True
        else:
            clustering_needed = False

        if clustering_needed:

            if len(clustering_list_col) == 0:
                clustering_list = clustering_list_row
                # print("no col clustering therefore, once again changing to row clustering list")

            # clustering based on the neighbouring pixels locations based on distance
            clustering_len_list = [clustering_list[0] + 1] # 1st element of the clustering len list

            for index_clus in range(len(clustering_list)): # any intermediate element of clustering len list
                if index_clus > 0:
                    clustering_len_list.append(clustering_list[index_clus] - clustering_list[index_clus-1])

            if len(clustering_list) > 0: # last element of the clustering len list
                clustering_len_list.append(len(row_indexes) - (clustering_list[len(clustering_list)-1] + 1))

            clustering_len_list = [abs(clustering_len_list_elem) for clustering_len_list_elem in clustering_len_list]

            # print("clustering_len_list", clustering_len_list)
            # checking the min index of the clustering len index, please add the code based on the length value, if length ==2 then ignore, then go to the next length

            # index of the minimum element of clustering_len_list
            min_index_clus = clustering_len_list.index(min(clustering_len_list))



            if len(clustering_len_list) > 0: #Check if clustering_len_list has more than zero elements

                if len(clustering_len_list) != len(set(clustering_len_list)):
                    has_duplicates = True


                if has_duplicates:
                    # print("There are duplicates except 1 and 2")

                    # Initialize the current cluster id

                    clusters = []
                    clusters_3_list = []
                    clusters_3_size_list = []

                    and_edges_t4_t5_3 = and_edges_t4_t5

                    # Find contours
                    contours_3, hierarchy_3 = cv2.findContours(and_edges_t4_t5, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)

                    cv2.cvtColor(and_edges_t4_t5_3, cv2.COLOR_GRAY2BGR)
                    # Iterate through contours and check if they are closed
                    contour_3_ID = 0

                    for contour_3 in contours_3:

                        if not cv2.isContourConvex(contour_3):
                            length_3 = cv2.arcLength(contour_3, closed=False)
                            # print("Cluster_ID", contour_3_ID, ":", contour_3.tolist())
                            clusters_3_list.append(contour_3.tolist())
                            clusters_3_size_list.append(length_3)

                        contour_3_ID = contour_3_ID + 1

                    clusters = clusters_3_list
                    # print("clusters_3_size_list", clusters_3_size_list)



                    # considering the size of the contour then considering the distance

                    only_ones_zeroes = all(clus_3_ele == 0 or clus_3_ele == 1 for clus_3_ele in clusters_3_size_list) #list has only ones and zeroes nothing else

                    if (1 in clusters_3_size_list and 0 in clusters_3_size_list) or clusters_3_size_list.count(1) > 1: #if list has zero and one both then ignore it or 1 used more than 2 times then ignore it

                        if only_ones_zeroes==False:

                            temp_list_without_zero = [temp_elem for temp_elem in clusters_3_size_list if temp_elem != 0 and temp_elem != 1]
                            min_val_except_zero = min(temp_list_without_zero)
                        else:
                            temp_list_without_zero = [temp_elem for temp_elem in clusters_3_size_list if temp_elem != 0]
                            min_val_except_zero = min(temp_list_without_zero)

                    elif len(set(clusters_3_size_list)) == 1:
                        min_val_except_zero = clusters_3_size_list[0]

                    else: #if list has zero then ignore it
                        temp_list_without_zero = [temp_elem for temp_elem in clusters_3_size_list if temp_elem != 0]
                        min_val_except_zero = min(temp_list_without_zero)


                    # print("min_val_except_zero", min_val_except_zero)

                    if 1 <= min_val_except_zero < 21:
                        # print("minimum index from length")
                        min_siz_clus_3_list_ind = clusters_3_size_list.index(min_val_except_zero)
                        min_dis_for_dupli_ind = min_siz_clus_3_list_ind

                        end_time_size = timer()
                        elapsed_time_size = (end_time_size - start_time_sing_line)*1000
                        print("elapsed_time_size", elapsed_time_size)
                        line_type = "Using Size"
                        time_val = elapsed_time_size

                    else:

                        dis_lin_pix_to_contu_list = []

                        # now select the one unit of the plant

                        key_for_clus = intersected_masks_list_final[list_of_keys_t5_ind]

                        # print("Part of the plant", key_for_clus[1])

                        class_name_from_list_8 = list_of_keys[key_for_clus[0]]  # change unit

                        selected_list_8 = dic_for_selection[class_name_from_list_8]
                        result_with_one_8 = np.where(selected_list_8 == 1)
                        listOfCoordinates_8 = list(zip(result_with_one_8[0], result_with_one_8[1]))

                        for cord_8 in listOfCoordinates_8:
                            img_temp_8[cord_8] = 1

                        img_for_cv_temp_7 = img_temp_8.astype(np.uint8) * 255

                        lsd_unit_1 = cv2.createLineSegmentDetector(refine=1, scale=0.5, sigma_scale=0.6, quant=0.5,
                                                                   ang_th=5.5, log_eps=0, density_th=0.1)

                        lines_unit_1 = lsd_unit_1.detect(img_for_cv_temp_7)[0]

                        # if (len(lines_unit_1) == 0): # no line segments to detect then use the other part of the unit
                        #     class_name_from_list_8 = list_of_keys[key_for_clus[0]]  # change unit
                        #
                        #     selected_list_8 = dic_for_selection[class_name_from_list_8]
                        #     result_with_one_8 = np.where(selected_list_8 == 1)
                        #     listOfCoordinates_8 = list(zip(result_with_one_8[0], result_with_one_8[1]))
                        #
                        #     for cord_8 in listOfCoordinates_8:
                        #         img_temp_8[cord_8] = 1
                        #
                        #     img_for_cv_temp_7 = img_temp_8.astype(np.uint8) * 255
                        #
                        #     lsd_unit_1 = cv2.createLineSegmentDetector(refine=1, scale=0.5, sigma_scale=0.6, quant=0.5,
                        #                                                ang_th=5.5, log_eps=0, density_th=0.1)
                        #
                        #     lines_unit_1 = lsd_unit_1.detect(img_for_cv_temp_7)[0]


                        copy_selected_list_units_7 = img_for_cv_temp_7 * 0

                        drawn_img_unit_7 = lsd_unit_1.drawSegments(copy_selected_list_units_7, lines_unit_1)

                        # Create the kernel
                        kernel = np.ones((7, 7), dtype=np.uint8)

                        # Perform dilation
                        drawn_img_unit_7_dil = cv2.dilate(drawn_img_unit_7, kernel, iterations=2)
                        drawn_img_unit_7_dil = cv2.erode(drawn_img_unit_7_dil, kernel, iterations=2)

                        drawn_img_unit_gray_7 = cv2.cvtColor(drawn_img_unit_7, cv2.COLOR_BGR2GRAY)

                        # img_for_cv_temp_7_RGB = cv2.cvtColor(img_for_cv_temp_7, cv2.COLOR_GRAY2BGR)

                        binary_image_line_unit_7 = cv2.adaptiveThreshold(drawn_img_unit_gray_7, 255,
                                                                         cv2.ADAPTIVE_THRESH_MEAN_C, \
                                                                         cv2.THRESH_BINARY, 15, -2)

                        substracted_img = drawn_img_unit_7_dil - drawn_img_unit_7

                        substracted_img = cv2.medianBlur(substracted_img, 7)

                        kernel_sub = cv2.getStructuringElement(cv2.MORPH_RECT,
                                                               (3, 5))  # dilation has done to connect pieces of stem

                        substracted_img = cv2.morphologyEx(substracted_img, cv2.MORPH_DILATE, kernel, iterations=3)

                        substracted_img = cv2.cvtColor(substracted_img, cv2.COLOR_BGR2GRAY)

                        # substracted_img_2 = substracted_img

                        # Find contours
                        contours, hierarchy = cv2.findContours(substracted_img, cv2.RETR_LIST,
                                                               cv2.CHAIN_APPROX_SIMPLE)

                        single_contour = max(contours, key=cv2.contourArea)

                        # Calculate the moments of the contour and centre point
                        moments = cv2.moments(single_contour)

                        # Calculate the center of mass of the contour
                        center_x = int(moments['m10'] / moments['m00'])
                        center_y = int(moments['m01'] / moments['m00'])

                        min_dis_lin_pix_to_contu_list = []

                        for clus_elem in clusters:

                            # print("clus_elem", clus_elem)

                            for cluster_coor_list_elem in clus_elem:
                                dis_lin_pix_to_contu_list = []
                                pix_to_contu_ang_list = []

                                dis_lin_pix_to_contu = round(np.sqrt((center_y - cluster_coor_list_elem[0][1]) ** 2 + (
                                            center_x - cluster_coor_list_elem[0][0]) ** 2), 3)

                                dis_lin_pix_to_contu_list.append(dis_lin_pix_to_contu)
                                # pix_to_contu_ang_list.append(pix_to_contu_ang)

                            min_dis_lin_pix_to_contu_list.append(min(dis_lin_pix_to_contu_list))

                        # print("minimum index from distance")
                        min_dis_for_dupli_ind = min_dis_lin_pix_to_contu_list.index(min(min_dis_lin_pix_to_contu_list))

                        end_time_dis = timer()
                        elapsed_time_dis = (end_time_dis - start_time_sing_line)*1000
                        line_type = "Using Distance"
                        time_val = elapsed_time_dis
                        print("elapsed_time_dis", elapsed_time_dis)


                    temp_clus_coord_list =[]

                    for elem_clus in clusters[min_dis_for_dupli_ind]:
                        temp_clus_coord_list.append(elem_clus[0])


                    cluster_coor_list = temp_clus_coord_list

                    # print("min_dis_lin_pix_to_contu_list", min_dis_lin_pix_to_contu_list)
                    # print("minimum index:", min_dis_for_dupli_ind)
                    # print("cluster_coor_list", cluster_coor_list)

                    img_temp_8 = np.zeros(img_temp_shape, dtype=np.uint8)

                else:
                    # print("There are no duplicates except 1 and 2")

                    # Find contours
                    contours_4, hierarchy_4 = cv2.findContours(and_edges_t4_t5, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)

                    clusters_4_list = []
                    clusters_4_size_list = []

                    for contour_4 in contours_4:

                        if not cv2.isContourConvex(contour_4):

                            length_4 = cv2.arcLength(contour_4, closed=False)
                            clusters_4_list.append(contour_4.tolist())
                            clusters_4_size_list.append(length_4)

                    # considering the size of the contour

                    all_same_val = False

                    if (1 in clusters_4_size_list and 0 in clusters_4_size_list) or clusters_4_size_list.count(1) > 1 :  # if list has zero and one both then ignore it or 1 used more than 2 times then ignore it

                        temp_list_without_zero_4 = [temp_elem_4 for temp_elem_4 in clusters_4_size_list if temp_elem_4 != 0 and temp_elem_4 != 1]

                        if len(temp_list_without_zero_4) > 1:
                            min_val_except_zero_4 = min(temp_list_without_zero_4)
                        else:
                            min_val_except_zero_4 = 1

                    elif len(set(clusters_4_size_list)) == 1:# whole list with same value
                        min_val_except_zero_4 = clusters_4_size_list[0]
                        temp_list_without_zero_4 = clusters_4_size_list
                        all_same_val = True

                    elif len(clusters_4_size_list) == 1:
                        min_val_except_zero_4 = clusters_4_size_list[0]

                    else:  # if list has zero then ignore it
                        temp_list_without_zero_4 = [temp_elem_4 for temp_elem_4 in clusters_4_size_list if temp_elem_4 != 0]
                        min_val_except_zero_4 = min(temp_list_without_zero_4)

                    # print("clusters_4_size_list", clusters_4_size_list)
                    # print("min_val_except_zero_4", min_val_except_zero_4)

                    if min_val_except_zero_4 > 21:
                        min_siz_clus_4_list_ind = clusters_4_size_list.index(min(temp_list_without_zero_4))
                        min_dis_for_dupli_ind_4 = min_siz_clus_4_list_ind

                    elif min_val_except_zero_4 == 1:
                        min_dis_for_dupli_ind_4 = clusters_4_size_list.index(1)
                    else:
                        min_siz_clus_4_list_ind = clusters_4_size_list.index(min(temp_list_without_zero_4))
                        min_dis_for_dupli_ind_4 = min_siz_clus_4_list_ind


                    temp_clus_coord_list_4 = []

                    for elem_clus_4 in clusters_4_list[min_dis_for_dupli_ind_4]:
                        temp_clus_coord_list_4.append(elem_clus_4[0])

                    if all_same_val and len(clusters_4_size_list) > 1: # if the list has same value and length of the list is higher than 1 then use all the coordinates
                        cluster_coor_list = []
                        for row_ind_sel, row_val_sel in enumerate(row_indexes):
                            cluster_coor_list.append([col_indexes[row_ind_sel], row_val_sel])

                    else:
                        cluster_coor_list = temp_clus_coord_list_4

                    # print("cluster_coor_list", cluster_coor_list)

                    end_time_size = timer()
                    elapsed_time_size = (end_time_size - start_time_sing_line)*1000
                    print("elapsed_time_size", elapsed_time_size )
                    line_type = "Using Size"
                    time_val = elapsed_time_size



        else:

            cluster_coor_list = []

            # if clustering_type == "row":
            for row_ind_sel, row_val_sel in enumerate(row_indexes):
                cluster_coor_list.append([col_indexes[row_ind_sel], row_val_sel])


            # print("cluster_coor_list_not_needed", cluster_coor_list)

            end_time_sing_line = timer()
            elapsed_time_sing_lin = (end_time_sing_line - start_time_sing_line)*1000
            print("elapsed_time_sing_lin", elapsed_time_sing_lin)
            line_type = "Only Single Line"
            time_val = elapsed_time_sing_lin


        to_time_excel.append([img_no, list_of_keys_t5_ind, elapsed_time_prep, line_type, time_val])


f.close()
end_time_final = timer()
elapsed_time_final = (end_time_final - start_time_prep)*1000
print("elapsed_time_final", elapsed_time_final)
print("to_time_excel", to_time_excel)


from openpyxl import load_workbook
from gc import collect

# Load the Excel workbook
workbook = load_workbook('D:/Academic/MSc/Masters Project 2023/Masters-Project-2023/pythonProject/Results_2/time_results_2.xlsx')
sheet = workbook.active

# Define the columns in the sheet
columns = {
    'No.': 1,
    'ID': 2,
    'Initial Preparation Time': 3,
    'Only Single Line': 4,
    'Using Size': 5,
    'Using Distance': 6
}

# Find the next available row to insert the new data
row_num = sheet.max_row + 1

# Loop through each row of data in the list and insert it into the sheet
for row in to_time_excel:
    # Insert the values into the correct columns based on the line type
    if row[3] == 'Only Single Line':
        sheet.cell(row=row_num, column=columns['No.']).value = row[0]
        sheet.cell(row=row_num, column=columns['ID']).value = row[1]
        sheet.cell(row=row_num, column=columns['Initial Preparation Time']).value = row[2]
        sheet.cell(row=row_num, column=columns['Only Single Line']).value = row[4]
    elif row[3] == 'Using Size':
        sheet.cell(row=row_num, column=columns['No.']).value = row[0]
        sheet.cell(row=row_num, column=columns['ID']).value = row[1]
        sheet.cell(row=row_num, column=columns['Initial Preparation Time']).value = row[2]
        sheet.cell(row=row_num, column=columns['Using Size']).value = row[4]
    elif row[3] == 'Using Distance':
        sheet.cell(row=row_num, column=columns['No.']).value = row[0]
        sheet.cell(row=row_num, column=columns['ID']).value = row[1]
        sheet.cell(row=row_num, column=columns['Initial Preparation Time']).value = row[2]
        sheet.cell(row=row_num, column=columns['Using Distance']).value = row[4]

    # Increment the row number
    row_num += 1

# Save the changes to the workbook
workbook.save('D:/Academic/MSc/Masters Project 2023/Masters-Project-2023/pythonProject/Results_2/time_results_2.xlsx')
collect()